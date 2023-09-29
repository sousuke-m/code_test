import glob

filelist = glob.glob('./task2/data/*.xlsx')

import openpyxl

out_workbook = openpyxl.Workbook()
out_sheet = out_workbook.active
out_sheet.title = '月報一覧'

out_sheet['A1'].value = '報告日'
out_sheet['B1'].value = '支店'
out_sheet['C1'].value = '担当者'
out_sheet['D1'].value = '売上目標（千円）'
out_sheet['E1'].value = '売上実績（千円）'
out_sheet['F1'].value = '差異（千円）'
out_sheet['G1'].value = '達成率（％）'

for i, file in enumerate(filelist):
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook['業務報告書']

    report_date = sheet['X1'].value
    office = sheet['X2'].value
    person = sheet['X3'].value
    sales_goal = sheet['H9'].value
    sales_result = sheet['H10'].value
    sales_diff = sheet['H11'].value
    sales_percentage = sheet['H12'].value

    out_sheet.cell(row=i+2, column=1).value = report_date
    out_sheet.cell(row=i+2, column=2).value = office
    out_sheet.cell(row=i+2, column=3).value = person
    out_sheet.cell(row=i+2, column=4).value = sales_goal
    out_sheet.cell(row=i+2, column=5).value = sales_result
    out_sheet.cell(row=i+2, column=6).value = sales_diff
    out_sheet.cell(row=i+2, column=7).value = sales_percentage

    out_sheet.cell(row=i+2, column=7).number_format = '0%'
    out_sheet.cell(row=i+2, column=1).number_format = 'yyyy年mm月dd日'

out_filename = './task2/output/月報一覧_202001.xlsx'
out_workbook.save(out_filename)